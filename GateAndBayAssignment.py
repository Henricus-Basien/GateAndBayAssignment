# -*- coding: utf-8 -*-
'''
Created on Tuesday 20.11.2018
Copyright (©) Henricus N. Basien
Author: Henricus N. Basien, Pavel Volkov
Email: Henricus@Basien.de
'''

#****************************************************************************************************
# ToDo
#****************************************************************************************************

"""
Objectives:
    (V) Walking Distances
    (V) Bay Preference
    ( ) Towing
Constraints:
    (V) Bay Compliance
    (V) Bay Compatibility / Fueling
    (~) Night Stays
    (~) Domestic or International
    ( ) Adjacency
"""

#****************************************************************************************************
# Imports
#****************************************************************************************************

import os,sys
RootPath = os.path.split(__file__)[0]
sys.path.insert(0,os.path.join(RootPath,"Scripts"))

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# External
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#----------------------------------------
# System
#----------------------------------------

#--- System ---
from collections import OrderedDict
# from copy import copy

#--- (Date-)Time ---
from time import time  as getTime
from time import sleep as wait
from datetime import datetime
Now = datetime.now

#--- Write TempFile ---
from os import dup, dup2, close
# import tempfile

#----------------------------------------
# Project Specific
#----------------------------------------

#--- Optimization ---
import pulp

#--- Mathematics ---
import numpy as np

#--- Excel Interface ---
import openpyxl

#--- Plotting ---
import matplotlib.pyplot as plt

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Internal
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

from ScheduleCreator import ScheduleCreator
from ReferenceDate import GetDate

#****************************************************************************************************
# GateAndBayAssignment
#****************************************************************************************************

class GateAndBayAssignmentSolver(object):
    """docstring for GateAndBayAssignmentSolver"""

    #================================================================================
    # Initialization
    #================================================================================
    
    def __init__(self, Airport,Schedule=None,LP_Path="Temp",AutoRun=True):
        super(GateAndBayAssignmentSolver, self).__init__()

        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        # Set Settings
        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        
        self.Airport  = Airport
        self.Schedule = Schedule

        #--- LP ---
        self.LP_Path = os.path.realpath(LP_Path)

        self.RemoveInfeasibleVariables = True
        self.AddAdjacencyConstraints   = False # ToDo: Fix this, because it makes the Gate Assignment impossibly slow!!!

        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        # Schedule
        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        if self.Schedule is None:
            MaxNrAircraft  = self.Airport.NrBays
            self.Scheduler = ScheduleCreator(self.Airport,MaxNrOverlappingAircraft=MaxNrAircraft,ScheduleFolder="Schedules",AutoRun=False)
            self.Scheduler.Run(Visualize=True)
            self.Schedule  = self.Scheduler.Schedule
        self.Schedule_dict = OrderedDict()
        for aircraft in self.Schedule:
            self.Schedule_dict[aircraft.ID] = aircraft

        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        # Run
        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        
        if AutoRun:
            self.RunAll()

    #================================================================================
    # Run Assignment
    #================================================================================
    
    def RunAll(self,Show=False):

        t0 = getTime()

        plt.close('all')

        self.Run(Mode="Bay")  # ; self.BayAssignment  = copy(self.SlotAssignment)
        self.SetupGatePreferences()
        self.Run(Mode="Gate") # ; self.GateAssignment = copy(self.SlotAssignment)

        self.ExportAllToExcel()

        dt = getTime()-t0
        print "*"*100
        print ">"*3+" "+"Complete Gate&Bay-Assignment Problem solved @"+str(Now())+"\t in "+str(round(dt/60.,1))+" min"
        print "*"*100

        if Show: plt.show()

    def Run(self,Mode,PrintProblem=False):

        print "*"*100
        print "Running LP Solver - "+Mode+"Assignment"
        print "*"*100

        t0 = getTime()

        #----------------------------------------
        # Set Mode
        #----------------------------------------

        self.Mode = Mode
        if   self.Mode == "Bay":
            self.Slots = self.Airport.Bays
        elif self.Mode == "Gate":
            self.Slots = self.Airport.Gates
        else:
            print "WARNING: Mode '"+self.Mode+"' is unknown! Possible Modes: ['Bay','Gate']"
            return

        #----------------------------------------
        # Run Problem
        #----------------------------------------
    
        self.CreateLP()
        self.ReconstructLP()
        if PrintProblem: print self.lp_problem
        self.RunLP()
        self.ExportResult()
        self.ConvertResult()
        self.PrintResult()

        #----------------------------------------
        # Addendum
        #----------------------------------------
    
        dt = getTime()-t0
        print "="*80
        print "Complete "+self.Mode+"Assignment Problem solved @"+str(Now())+"\t in "+str(round(dt/60.,1))+" min"
        print "="*80

        self.PlotResult()

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Create
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def CreateLP(self):

        t0 = getTime()

        print "> Problem is being created..."

        LP = []

        if self.Mode=="Bay":
            self.InFeasibleVariables = self.GetLP_BayCompatibilityConstraints()
        elif self.Mode=="Gate":
            self.InFeasibleVariables = []
        #--- Variables & Objectives ---
        LP+=self.GetLP_Variables()
        if self.Mode=="Gate":
            self.SetAdjacencyConstraints()
            LP+=self.GetLP_AdjacencyVariables()
        LP+=self.GetLP_ObjectiveFunctions()
        #--- Constaints ---
        if not self.RemoveInfeasibleVariables:
            LP+=self.InFeasibleVariables
        LP+=self.GetLP_SlotComplianceConstraints()
        LP+=self.GetLP_TimeConstraints()

        #..............................
        # Write to File
        #..............................
        
        if not os.path.exists(self.LP_Path): os.makedirs(self.LP_Path)
        filename = self.FormatTitle("LP.PuLP")
        self.LP_filepath = os.path.join(self.LP_Path,filename)
        NrLines = 0
        with open(self.LP_filepath,"w") as LP_File:
            for line in LP:
                if line[0]!="#":NrLines+=1
                LP_File.write(line+"\n")

        print str(NrLines)+" Variables/ObjectiveFunctions/Constraints have been created!"

        dt = getTime()-t0
        print ">"*3+" "+"Problem created       @"+str(Now())+"\t in "+str(round(dt,2))+"s"

    #----------------------------------------
    # Variables
    #----------------------------------------

    def GetLP_Variables(self):

        self.Variables = []
        
        for i in range(len(self.Schedule)):
            for k in range(len(self.Slots)):
                var = self.GetVarName(i,k)
                if self.RemoveInfeasibleVariables and var in self.InFeasibleVariables: continue
                self.Variables.append(var)
        
        Variables = ["Var: "+v for v in self.Variables]

        Variables.append("#"*100)

        return Variables

    def GetLP_AdjacencyVariables(self):

        if not self.AddAdjacencyConstraints:
            return []

        self.AdjacencyVariables = self.AdjacencyConstraints.values()

        Variables = ["Var: "+v for v in self.AdjacencyVariables]

        Variables.append("#"*100)

        return Variables

    def GetVarName(self,i,k):
        return "X_"+str(i)+"_"+str(k)
        # return self.Schedule[i].Name+"_"+self.Slots[k].Name

    #----------------------------------------
    # Objective Function(s)
    #----------------------------------------
    
    def GetLP_ObjectiveFunctions(self):

        if   self.Mode=="Bay":
            return self.GetLP_BayObjectiveFunctions()
        elif self.Mode=="Gate":
            return self.GetLP_GateObjectiveFunctions()

    def GetLP_BayObjectiveFunctions(self):
        # Objective function
        ObjectiveFunctions = []

        #..............................
        # Objective TransportDistance (Z1)
        #..............................
    
        ObjectiveFunctions.append("# ObjectiveFunction - TransportDistance")
        ObjectiveFunctions.append("Var: Z1")

        ObjectiveFunction_TransportDistance = ""
        for i in range(len(self.Schedule)):
            a = self.Schedule[i]
            for k in range(len(self.Airport.Bays)):
                b = self.Airport.Bays[k]
                var = self.GetVarName(i,k)
                if self.RemoveInfeasibleVariables and var in self.InFeasibleVariables: continue
                #--- Objective ---
                if b.Virtual:
                    TravelDistance = 10**6
                    #.. Try to stack virtual elements ..
                    try:
                        index = int(b.Name[1:])
                        TravelDistance*=(index+1)
                    except: print "WARNING: Unable to get Virtual Bay Index!"
                else:
                    Terminal = a.Airline.Terminal
                    BayName = b.Name
                    if not (Terminal,BayName) in self.Airport.TravelDistances_Bays.keys():
                        BayName = BayName.rstrip("ABCD LR")
                    TravelDistance = self.Airport.TravelDistances_Bays[(Terminal,BayName)]
                ObjectiveFunction_TransportDistance+= var+"*"+str(a.NrPassengers)+"*"+str(TravelDistance)+ "+"
        ObjectiveFunction_TransportDistance+=' - Z1 == 0'

        ObjectiveFunctions.append(ObjectiveFunction_TransportDistance)

        #..............................
        # Objective AirlinePreference (Z2)
        #..............................
    
        ObjectiveFunctions.append("# ObjectiveFunction - AirlinePreference")
        ObjectiveFunctions.append("Var: Z2")

        ObjectiveFunction_AirlinePreference = ""
        for i in range(len(self.Schedule)):
            a = self.Schedule[i]
            for k in range(len(self.Airport.Bays)):
                var = self.GetVarName(i,k)
                if self.RemoveInfeasibleVariables and var in self.InFeasibleVariables: continue
                b = self.Airport.Bays[k]
                if hasattr(a,"BayPreference") and a.BayPreference is not None and a.BayPreference==b.Name:
                    ObjectiveFunction_AirlinePreference+= var+ "+"
        ObjectiveFunction_AirlinePreference+=' - Z2 == 0'

        ObjectiveFunctions.append(ObjectiveFunction_AirlinePreference)

        #..............................
        # Objective RelocationPenalty (Z3)
        #..............................
    
        ObjectiveFunctions.append("# ObjectiveFunction - RelocationPenalty")
        ObjectiveFunctions.append("Var: Z3")

        ObjectiveFunction_RelocationPenalty = ""
        ObjectiveFunction_RelocationPenalty+=' - Z3 == 0'

        ObjectiveFunctions.append(ObjectiveFunction_RelocationPenalty)

        #..............................
        # Combine Objective
        #..............................
        
        ObjectiveFunctions.append("# ObjectiveFunction - Combined")

        Z1_Max = np.max([a.NrPassengers for a in self.Schedule])*np.max(self.Airport.TravelDistances_Bays.values()) #1

        alpha = 1.0
        beta  = Z1_Max
        gamma = 3*beta

        if 1:
            ObjectiveFunctions.append("Var: Z4")
            ObjectiveFunctions.append(str(alpha)+'*Z1 - '+str(beta)+'*Z2 - Z4 ==0')
            ObjectiveFunctions.append("Var: Z5")
            ObjectiveFunctions.append('Z4 + '+str(gamma)+'*Z3 - Z5 ==0')      
            ObjectiveFunctions.append('Z5 , "Z"')        
        
        else:
            ObjectiveFunctions.append(str(alpha)+'*Z1 - '+str(beta)+'*Z2 + '+str(gamma)+'*Z3 , "Z"')        

        #..............................
        # Addendum
        #..............................
        
        ObjectiveFunctions.append("#"*100)

        return ObjectiveFunctions

    def GetLP_GateObjectiveFunctions(self):
        # Objective function
        ObjectiveFunctions = []

        #..............................
        # Objective TransportDistance (Z6)
        #..............................
    
        ObjectiveFunctions.append("# ObjectiveFunction - TransportDistance")
        ObjectiveFunctions.append("Var: Z6")

        ObjectiveFunction_TransportDistance = ""
        for i in range(len(self.Schedule)):
            a = self.Schedule[i]
            for k in range(len(self.Airport.Gates)):
                g = self.Airport.Gates[k]
                var = self.GetVarName(i,k)
                if self.RemoveInfeasibleVariables and var in self.InFeasibleVariables: continue
                #--- Objective ---
                if g.Virtual:
                    TravelDistance = 10**6
                    #.. Try to stack virtual elements ..
                    try:
                        index = int(g.Name[1:])
                        TravelDistance*=(index+1)
                    except: print "WARNING: Unable to get Virtual Gate Index!"
                else:
                    Terminal = a.Airline.Terminal
                    GateName = g.Name
                    if not (Terminal,GateName) in self.Airport.TravelDistances_Gates.keys():
                        GateName = GateName.rstrip("ABCD LR")
                    TravelDistance = self.Airport.TravelDistances_Gates[(Terminal,GateName)]
                ObjectiveFunction_TransportDistance+= var+"*"+str(a.NrPassengers)+"*"+str(TravelDistance)+ "+"
        ObjectiveFunction_TransportDistance+=' - Z6 == 0'

        ObjectiveFunctions.append(ObjectiveFunction_TransportDistance)

        #..............................
        # Objective AirlinePreference (Z7)
        #..............................
    
        ObjectiveFunctions.append("# ObjectiveFunction - AirlinePreference")
        ObjectiveFunctions.append("Var: Z7")

        ObjectiveFunction_AirlinePreference = ""
        for i in range(len(self.Schedule)):
            a = self.Schedule[i]
            for k in range(len(self.Airport.Gates)):
                var = self.GetVarName(i,k)
                if self.RemoveInfeasibleVariables and var in self.InFeasibleVariables: continue
                g = self.Airport.Gates[k]
                if hasattr(a,"GatePreference") and a.GatePreference is not None and a.GatePreference==g.Name:
                    ObjectiveFunction_AirlinePreference+= var+ "+"
        ObjectiveFunction_AirlinePreference+=' - Z7 == 0'

        ObjectiveFunctions.append(ObjectiveFunction_AirlinePreference)

        #..............................
        # Objective Adjacency (Z8)
        #..............................
    
        ObjectiveFunctions.append("# ObjectiveFunction - Adjacency")
        ObjectiveFunctions.append("Var: Z8")

        ObjectiveFunction_Adjacency = ""
        for var in self.AdjacencyConstraints.values():
            ObjectiveFunction_Adjacency+=var+" + "
        ObjectiveFunction_Adjacency+=' - Z8 == 0'

        ObjectiveFunctions.append(ObjectiveFunction_Adjacency)
        
        #..............................
        # Combine Objective
        #..............................

        Z6_Max = np.max([a.NrPassengers for a in self.Schedule])*np.max(self.Airport.TravelDistances_Gates.values()) #1

        delta   = 1.0
        epsilon = Z6_Max
        zeta    = Z6_Max+2*epsilon

        ObjectiveFunctions.append(str(delta)+'*Z6 - '+str(epsilon)+'*Z7 + '+str(zeta)+'*Z8 , "Z"')       

        #..............................
        # Addendum
        #..............................
        
        ObjectiveFunctions.append("#"*100)

        return ObjectiveFunctions

    #----------------------------------------
    # Constraints
    #----------------------------------------

    #..............................
    # Slot Compliance
    #..............................
    
    def GetLP_SlotComplianceConstraints(self):

        SlotComplianceConstraints = []

        SlotComplianceConstraints.append("# SlotCompliance Constraints")

        for i in range(len(self.Schedule)):
            SlotComplianceConstraint = ""
            for k in range(len(self.Slots)):
                var = self.GetVarName(i,k)
                if self.RemoveInfeasibleVariables and var in self.InFeasibleVariables: continue
                SlotComplianceConstraint+=var + " + "
            if SlotComplianceConstraint=="": continue # Avoids Empty Constraints

            SlotComplianceConstraint+=" 0 == 1"
            SlotComplianceConstraints.append(SlotComplianceConstraint)

        return SlotComplianceConstraints     

    #..............................
    # Time Constraints
    #..............................

    def GetLP_TimeConstraints(self):

        TimeConstraints = []

        TimeConstraints.append("# Time Constraints")

        for i in range(len(self.Schedule)):
            a1 = self.Schedule[i]
            for j in range(i,len(self.Schedule)):
                a2 = self.Schedule[j]
                #--- Avoid unrequired Constraints ---
                if i==j: continue
                if not self.ScheduleClash(a1,a2): continue
                #--- Cycle all Slots ---
                for k in range(len(self.Slots)):
                    var1 = self.GetVarName(i,k)
                    var2 = self.GetVarName(j,k)
                    if self.RemoveInfeasibleVariables and (var1 in self.InFeasibleVariables or var2 in self.InFeasibleVariables): continue
                    TimeConstraint = var1 + " + " + var2
                    AdjacencyKey = (i,j)#(k,i,j)
                    if self.Mode=="Gate" and AdjacencyKey in self.AdjacencyConstraints:
                        TimeConstraint+=" - "+self.AdjacencyConstraints[AdjacencyKey]
                    TimeConstraint+=" <= 1"
                    TimeConstraints.append(TimeConstraint)

        return TimeConstraints

    def SetAdjacencyConstraints(self):

        self.AdjacencyConstraints = OrderedDict()

        if not self.AddAdjacencyConstraints:
            return

        for i in range(len(self.Schedule)):
            a1 = self.Schedule[i]
            for j in range(i,len(self.Schedule)):
                a2 = self.Schedule[j]
                #--- Avoid unrequired Constraints ---
                if i==j: continue
                if not self.ScheduleClash(a1,a2): continue
                #--- Set Adjacancy ---
                AdjacencyKey = (i,j)
                self.AdjacencyConstraints[AdjacencyKey] = "S_"+str(i)+"_"+str(j)
                ##--- Cycle all Slots ---
                # for k in range(len(self.Slots)):
                #     AdjacencyKey = (k,i,j)
                #     self.AdjacencyConstraints[AdjacencyKey] = "S_"+str(k)+"_"+str(i)+"_"+str(j)

    def ScheduleClash(self,a1,a2):
        #--- Base Constraints ---
        if   a1.Arrival  <a2.Departure and a1.Arrival  >a2.Arrival: return True
        elif a1.Departure<a2.Departure and a1.Departure>a2.Arrival: return True
        #--- Mirrored Constraints ---
        elif a2.Arrival  <a1.Departure and a2.Arrival  >a1.Arrival: return True
        elif a2.Departure<a1.Departure and a2.Departure>a1.Arrival: return True

        return False

    #----------------------------------------
    # Bay Compatibility
    #----------------------------------------
    
    def GetLP_BayCompatibilityConstraints(self):

        BayCompatibilityConstraints = []

        BayCompatibilityConstraints.append("# BayCompatibility Constraints")

        for i in range(len(self.Schedule)):
            a = self.Schedule[i]
            #--- Cycle all Bays ---
            for k in range(len(self.Airport.Bays)):
                b = self.Airport.Bays[k]
                if (b.CompatibleAircraftTypes is not None and a.Type not in b.CompatibleAircraftTypes) or (a.NeedsFueling and not b.FuelingPossible):
                    BayCompatibilityConstraints.append(self.GetVarName(i,k)+" == 0")

        #--- Extract Variables ---
        if self.RemoveInfeasibleVariables:
            for i in range(len(BayCompatibilityConstraints))[::-1]:
                c = BayCompatibilityConstraints[i]
                if c[0]=="#": del BayCompatibilityConstraints[i]
                else: BayCompatibilityConstraints[i] = c.split(" ")[0]

            if 1:
                TotalNrVariables = len(self.Schedule)*len(self.Airport.Bays)
                InFeasiblePer = len(BayCompatibilityConstraints)/float(TotalNrVariables)
                #print BayCompatibilityConstraints
                print ">"*2+" "+str(len(BayCompatibilityConstraints))+"/"+str(TotalNrVariables)+" ("+str(round(InFeasiblePer*100,1))+r"%) InFeasibleVariables Found!"

        return BayCompatibilityConstraints

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Reconstruct
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def ReconstructLP(self):

        t0 = getTime()

        print "> Problem is being reconstructed..."

        # print "Init Problem"
        self.lp_problem = pulp.LpProblem(self.Mode+" Assignment Problem", pulp.LpMinimize)

        with open(self.LP_filepath) as LP_File:

            lines = LP_File.readlines()
            LineNr = 0
            for line in lines:
                LineNr+=1
                line = line.strip()
                if line=="" or line[0]=="#": continue

                # print "Adding: ",LineNr#,line
                # sys.stdout.write(str(LineNr)+",")
                #--- Add Variables ---
                if line[:4]=="Var:":
                    var = line[5:]
                    if var[0]=="Z":
                        VarType = "Continuous"
                    else:
                        VarType = "Binary"
                    exec(var+" = pulp.LpVariable('"+var+"', cat='"+VarType+"')")
                    continue
                #--- Add Constraints ---
                exec("self.lp_problem+="+line)

        dt = getTime()-t0
        print ">"*3+" "+"Problem reconstructed @"+str(Now())+"\t in "+str(round(dt,2))+"s"

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Run 
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def RunLP(self,time_limit = 8*3600): #1*3600):

        t0 = getTime()

        print "> Problem is being solved..."

        solver = pulp.PULP_CBC_CMD(msg=True, maxSeconds=time_limit) # None

        if 0: #1:
            self.lp_problem.solve(solver=solver)
        else:

            #----------------------------------------
            # Write/Solve
            #----------------------------------------
                
            filename = self.FormatTitle("LP_Output.txt")
            with open(os.path.join(self.LP_Path,filename),"w") as tmp_output: #tempfile.TemporaryFile() as tmp_output:
                orig_std_out = dup(1)
                dup2(tmp_output.fileno(), 1)
                result = self.lp_problem.solve(solver=solver)
                #print result
                dup2(orig_std_out, 1)
                close(orig_std_out)
                tmp_output.seek(0)

            #----------------------------------------
            # Read
            #----------------------------------------
            
            if 0:
                def Print(text):
                    print text
                
                with open(os.path.join(self.LP_Path,filename),"r") as tmp_output: 
                    lines = tmp_output.read().splitlines()
                    [Print(line.decode('ascii')) for line in lines]
                

        dt = getTime()-t0
        print ">"*3+" "+"Problem solved        @"+str(Now())+"\t in "+str(round(dt,2))+"s"

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Print
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def PrintResult(self,ShowVariables=False):

        print "*"*100

        #----------------------------------------
        # Variables
        #----------------------------------------
        
        if ShowVariables:
            print "Variables:"

            for variable in self.lp_problem.variables():
                print " "*3+"{} = {}".format(variable.name, variable.varValue)

        #----------------------------------------
        # Status
        #----------------------------------------
        
        print "Status: "+str(pulp.LpStatus[self.lp_problem.status])

        #----------------------------------------
        # Objective
        #----------------------------------------
        
        print "Objective:"
        Objective = self.lp_problem.objective
        if Objective is not None:
            Objective = pulp.value(Objective)
        print " "*3+str(Objective)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Convert Result
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        
    def ConvertResult(self,PrintResult=False):#True):

        self.SlotAssignment      = OrderedDict()
        self.SlotAssignment_Dual = OrderedDict()

        for variable in self.lp_problem.variables():
            if variable.varValue==0: continue
            if variable.name[0]=="Z": continue
            try:
                X,i,k = variable.name.split("_")
            except:
                print "WARNING: Unable to parse Variable '"+variable.name+"'"
                continue

            if X!="X":
                continue

            i = int(i);k = int(k)

            K = str(k)
            self.SlotAssignment[self.Schedule[i].ID] = self.Slots[k]
            #--- Dual ---
            if not K in self.SlotAssignment_Dual:
                self.SlotAssignment_Dual[K] = []
            self.SlotAssignment_Dual[K].append(self.Schedule[i])

            #--- Store in Aircraft ---
            aircraft = self.Schedule[i]
            if   self.Mode=="Bay":
                aircraft.BayAssigned  = self.Slots[k].Name
            elif self.Mode=="Gate":
                aircraft.GateAssigned = self.Slots[k].Name

        #--- Sort ---
        self.SlotAssignment_Dual = OrderedDict(sorted(self.SlotAssignment_Dual.items(), key=lambda t: t[0]))

        #--- Print ---
        if PrintResult:
            print "FlightID | "+self.Mode+"ID"
            for key in self.SlotAssignment.keys():
                print key,"\t|",self.SlotAssignment[key]

    #----------------------------------------
    # Gate Preferences
    #----------------------------------------
    
    def SetupGatePreferences(self):

        for FlightID in self.SlotAssignment:
            aircraft = self.Schedule_dict[FlightID]
            bay      = self.SlotAssignment[FlightID]
            if (aircraft.GatePreference is None) and (bay.Name in list(self.Airport.Gates_dict.keys())):
                gate = self.Airport.Gates_dict[bay.Name]
                if not gate.Virtual:
                    aircraft.GatePreference = bay.Name

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Export
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def ExportResult(self):

        filename = self.FormatTitle("LP_Result.txt")
        with open(os.path.join(self.LP_Path,filename),"w") as ResultsFile:
            for variable in self.lp_problem.variables():
                ResultsFile.write("{} = {}".format(variable.name, variable.varValue)+"\n")

    def ExportAllToExcel(self):

        wb = openpyxl.Workbook()
        ws = wb.active

        #----------------------------------------
        # Write Header
        #----------------------------------------
        
        #--- Info ---
        ws.cell(row=1, column=1 ).value = "ID"
        ws.cell(row=1, column=2 ).value = "Type"
        ws.cell(row=1, column=3 ).value = "Airline"
        ws.cell(row=1, column=4 ).value = "Arrival [datetime]"
        ws.cell(row=1, column=5 ).value = "Departure [datetime]"
        #--- Assignment ---
        ws.cell(row=1, column=6 ).value = "Bay-Assigned"
        ws.cell(row=1, column=7 ).value = "Bay-Prefered"
        ws.cell(row=1, column=8 ).value = "Bay-PreferenceMet"
        ws.cell(row=1, column=9 ).value = "Gate-Assigned"
        ws.cell(row=1, column=10).value = "Gate-Prefered"
        ws.cell(row=1, column=11).value = "Gate-PreferenceMet"

        #----------------------------------------
        # Write Data
        #----------------------------------------

        for i in range(len(self.Schedule)):
            aircraft = self.Schedule[i]
            #--- Info ---
            ws.cell(row=i+2, column=1 ).value = aircraft.ID
            ws.cell(row=i+2, column=2 ).value = aircraft.Type
            ws.cell(row=i+2, column=3 ).value = aircraft.Airline.Name
            ws.cell(row=i+2, column=4 ).value = aircraft.Arrival#.time()
            ws.cell(row=i+2, column=5 ).value = aircraft.Departure#.time()
            #--- Assignment ---
            ws.cell(row=i+2, column=6 ).value  = aircraft.BayAssigned
            ws.cell(row=i+2, column=7 ).value  = aircraft.BayPreference
            ws.cell(row=i+2, column=8 ).value  = aircraft.BayAssigned == aircraft.BayPreference
            ws.cell(row=i+2, column=9 ).value  = aircraft.GateAssigned
            ws.cell(row=i+2, column=10).value  = aircraft.GatePreference
            ws.cell(row=i+2, column=11).value  = aircraft.GateAssigned == aircraft.GatePreference

        title = self.Scheduler.FormatTitle("Gate&BayAssignment Schedule.xlsx")
        wb.save(os.path.join(self.Scheduler.ScheduleFolder,title))
        print ">"+" "+"Exported '"+title+"'"

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Plot Result
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def PlotResult(self,fontsize=4,Show=False):

        # if Show:
        #     plt.close('all')

        fig,ax=plt.subplots(figsize=(16,9),dpi=self.Scheduler.dpi)

        #----------------------------------------
        # Preprocess
        #----------------------------------------
        
        # MinDate = ReferenceDate
        MinDate = GetDate(np.min([a.Arrival for a in self.Schedule]))

        #----------------------------------------
        # Set Data
        #----------------------------------------
        
        labels=[]
        for k in range(len(self.Slots)):
            #--- Add Labels ---
            Slot = self.Slots[k]
            labels.append(str(Slot))
            #--- Check Use ---
            K = str(k)
            if not K in self.SlotAssignment_Dual.keys(): continue
            for aircraft in self.SlotAssignment_Dual[K]:
                color = Slot.Color #None
                t_a = (aircraft.Arrival  -MinDate).total_seconds()/3600.
                t_d = (aircraft.Departure-MinDate).total_seconds()/3600.
                dt = t_d-t_a
                ax.broken_barh([(t_a,dt)], (k-0.4,0.8),color=color)

                if 1:
                    ax.annotate(aircraft.Type+" - ID: "+aircraft.ID+"\n"+"> #P: "+str(aircraft.NrPassengers), (t_a,k),
                    fontsize=fontsize,
                    horizontalalignment='left', verticalalignment='center')

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        self.Scheduler.ShowAirportDayLines()

        #..............................
        # Labels
        #..............................
        
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=fontsize) 
        ax.set_xlabel("Time [h]")

        #..............................
        # Layout
        #..............................
        
        plt.xlim(left=0)
        plt.tight_layout()
        title=self.FormatTitle()
        plt.savefig(os.path.join(self.Scheduler.ScheduleFolder,title))
        if Show:
            plt.show(title)

    #================================================================================
    # Formatting
    #================================================================================
    
    def FormatTitle(self,title=""):

        return (self.Scheduler.FormatTitle(self.Mode+"Assignment"+" "+title)).strip()

#****************************************************************************************************
# Test Code
#****************************************************************************************************

def SolveGateAndBayAssignmentProblem(Seed=None,OnlyCreateSchedule=False):
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Seed
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    if Seed is None:
        Seed = raw_input(">>> Please Enter a seed for the stochastic Realization [ENTER for random value]: ")
        Seed = Seed.strip()
        if Seed=="":
            Seed = np.random.randint(10**9)
    try:
        Seed = int(Seed)
    except:
        print "ERROR: Seed '"+str(Seed)+"' is not possible, numerical values only!"
        quit()

    np.random.seed_backup = Seed
    np.random.seed(Seed)

    print "*"*100
    print ">"*3+" "+"Initialized Simulation with Seed '"+str(Seed)+"'"
    print "*"*100

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Imports
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    from JKIA import JKIA

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Run Test
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    airport = JKIA()
    airport.PrintInfo()

    GABA_Solver = GateAndBayAssignmentSolver(airport,AutoRun = not OnlyCreateSchedule)

#----------------------------------------
# SaveGuard
#----------------------------------------

def SaveGuard(func,*args,**kwargs):
    try:
        func(*args,**kwargs)
    except:
        import traceback
        print traceback.format_exc()

def SaveSolver(*args,**kwargs):
    from GateAndBayAssignment import SolveGateAndBayAssignmentProblem,SaveGuard
    
    SaveGuard(SolveGateAndBayAssignmentProblem,*args,**kwargs)

if __name__=="__main__":

    print "GateAndBayAssignmentSolver Started on: "+str(Now())

    PossibleModes = ["Manual","BatchScheduler","MultiSolver"]
    RunMode = raw_input("Please select GateAndBayAssignmentSolver Mode "+str(PossibleModes)+": ")
    RunMode = RunMode.strip()

    try:
        RunMode = int(RunMode)
        RunMode = PossibleModes[RunMode-1]
    except:
        pass

    if RunMode=="":
        RunMode = "Manual"

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Manual Mode
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    if RunMode.lower()=="manual":#1:
        Seed = None # np.random.random()
        SolveGateAndBayAssignmentProblem(Seed)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # MultiProcessed Solver
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    elif RunMode.lower()=="multisolver" or RunMode.lower()=="batchscheduler":
        import pp,multiprocessing
        # from functools import partial
        nCPU = int(multiprocessing.cpu_count()/2-1)
        PP_Server = pp.Server(nCPU,restart=True)

        print "Started PP Server with '"+str(nCPU)+"' CPUs"

        #----------------------------------------
        # Settings
        #----------------------------------------
        
        if RunMode.lower()=="batchscheduler": OnlyCreateSchedule = True
        else:                                 OnlyCreateSchedule = False

        #----------------------------------------
        # Get Seeds
        #----------------------------------------
        
        if RunMode.lower()=="batchscheduler":
            Seeds = range(100)
        else:
            print "Please enter the Seed Values you'd like to run [ENTER to exit | 'r' for random]:"
            Seeds = []
            while True:
                Seed = raw_input("SeedValue #"+str(len(Seeds)+1)+": ")
                Seed = Seed.strip()
                if Seed=="":
                    break
                elif Seed=="r":
                    Seed = np.random.randint(10**9)
                try:
                    Seed = int(Seed)
                except:
                    print "WARNING: Seed '"+str(Seed)+"' is not valid; only use numerical values!"
                    continue
                Seeds.append(Seed)

        #----------------------------------------
        # Create Jobs
        #----------------------------------------
        
        Jobs = []
        for i in range(len(Seeds)):
            Seed = Seeds[i]
            print "Creating new PP_Job #"+str(i+1)+": Seed="+str(Seed)
            cmd    = SaveSolver #SolveGateAndBayAssignmentProblem
            params = (Seed,OnlyCreateSchedule)

            job = PP_Server.submit(cmd,params)
            Jobs.append(job)#partial(SaveGuard,cmd),params))
            job.DONE = False

        #----------------------------------------
        # Run Jobs
        #----------------------------------------
        
        print ">"*3+" Running "+str(len(Seeds))+" Jobs..."
        Results = []
        if 0:#1:
            for i,job in enumerate(Jobs):
                Seed = Seeds[i]
                print "Waiting for PP_Job #"+str(i+1)+": Seed="+str(Seed)
                Result = job()
                Results.append(Result)
        while len(Jobs)!=len(Results):
            for i,job in enumerate(Jobs):#([j for j in Jobs if (j.finished and not j.DONE)]):
                if not (job.finished and not job.DONE): continue
                Seed = Seeds[i]
                print ">"*50+"PP_Job #"+str(i+1)+" ("+str(len(Results))+")"+" - Finished! : Seed="+str(Seed)
                Result = job()
                Results.append(Result)
                job.DONE = True
            wait(1.0)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Unknown
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    else:
        print "ERROR: Mode '"+str(RunMode)+"' is unknown! Please select one of the PossibleModes: "+str(PossibleModes)
