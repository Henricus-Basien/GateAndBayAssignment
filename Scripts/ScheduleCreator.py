# -*- coding: utf-8 -*-
'''
Created on Monday 19.11.2018
Copyright (©) Henricus N. Basien
Author: Henricus N. Basien
Email: Henricus@Basien.de
'''

#****************************************************************************************************
# Imports
#****************************************************************************************************

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# External
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#----------------------------------------
# System
#----------------------------------------

#--- System ---
import os
from copy import copy
from collections import OrderedDict,Counter

#--- (Date-)Time ---
from time import time as getTime
import datetime
Now = datetime.datetime.now

#----------------------------------------
# Project Specific
#----------------------------------------

#--- Mathematics ---
import numpy as np

#--- Excel Interface ---
import openpyxl

#--- Plotting ---
import matplotlib.pyplot as plt

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Internal
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

from ReferenceDate import ReferenceDate,GetDate

def getRoundedThreshold(a, MinClip,SetInt=False):
    frac = float(a) / MinClip
    if SetInt:
        frac = int(frac)
    return np.round(frac) * MinClip

dt0 = datetime.datetime(year=1900,month=1,day=1)

#****************************************************************************************************
# ScheduleCreator
#****************************************************************************************************

class ScheduleCreator(object):
    """docstring for ScheduleCreator"""

    #================================================================================
    # Initialization
    #================================================================================
    
    def __init__(self, Airport,MaxNrAircraft=None,MaxNrOverlappingAircraft=30,MaxNrDays=1,ScheduleFolder="Temp",AutoRun=True):
        super(ScheduleCreator, self).__init__()

        self.Airport                  = Airport
        self.MaxNrAircraft            = MaxNrAircraft
        self.MaxNrOverlappingAircraft = MaxNrOverlappingAircraft
        self.MaxNrDays                = MaxNrDays

        self.ScheduleFolder = os.path.realpath(ScheduleFolder)
        if not os.path.exists(self.ScheduleFolder): os.makedirs(self.ScheduleFolder)

        if self.MaxNrAircraft is None:
            AverageStayTime = 2.75#3.5#3.0#2.5 # [h]
            TimeFactor = AverageStayTime * 2
            self.MaxNrAircraft = int((self.Airport.GetOperationalTime()/3600.)*self.MaxNrOverlappingAircraft/float(TimeFactor))*self.MaxNrDays

        self.dpi = 300#120

        if AutoRun:
            self.Run()
        
    #================================================================================
    # Schedule Creator
    #================================================================================

    def Run(self,Print=False,Export=True,Visualize=True):

        t0 = getTime()

        self.CreateAircraftSchedule()
        if Print:     self.PrintSchedule()
        if Export:    self.ExportScheduleToExcel()
        if Visualize: self.Visualize()

        dt = getTime()-t0
        print "="*80
        print "Created Schedule ("+str(len(self.Schedule))+"-Aircraft) @"+str(Now())+"\t in "+str(round(dt,1))+"s"
        print "="*80

    def CreateAircraftSchedule(self,TimeRange=[45*60,20*3600],MinTimeStep=5*60,Sort=True,RefDate=ReferenceDate,NightStayMandatory=True):

        self.Schedule = []

        BetaMean = np.mean([TimeRange[0]]*20+[TimeRange[1]])
        # print "BetaMean",BetaMean/3600.,"h"

        for i in range(self.MaxNrAircraft):

            #----------------------------------------
            # Airline
            #----------------------------------------

            if self.Airport.LocalAirline is None or float(i)/self.MaxNrAircraft>0.35:
                airline = np.random.choice(self.Airport.Airlines)
            else:
                airline = copy(self.Airport.LocalAirline)

            #----------------------------------------
            # AircraftType
            #----------------------------------------
            
            AircraftType = None
            while AircraftType is None or AircraftType.Type not in self.Airport.CompatibleAircraftTypes:
                AircraftType = np.random.choice(airline.AircraftTypes)
            ID = i+1

            ID = airline.Name+str(ID)

            #----------------------------------------
            # Arrival/Departure
            #----------------------------------------

            Open_s   = (self.Airport.T_Open-dt0).total_seconds()
            Close_s  = (self.Airport.T_Close-dt0).total_seconds()
            Opened_s = (self.Airport.T_Close-self.Airport.T_Open).total_seconds()

            #--- Arrival ---
            arrivalper = np.random.uniform()
            if 1:
                arrivalper = (arrivalper-0.5)*2 # Shift to [-1,+1]
                arrivalper = abs(arrivalper)**(1./1.4) * (abs(arrivalper)/arrivalper)      # Modulate
                arrivalper = (arrivalper+1)/2   # Reshift to [0,+1]
            ArrivalT   = Open_s + arrivalper * Opened_s # np.random.uniform((self.Airport.T_Open-dt0).total_seconds(),(self.Airport.T_Close-dt0).total_seconds())
            if MinTimeStep is not None: ArrivalT = getRoundedThreshold(ArrivalT,MinTimeStep)
            Arrival   = RefDate + datetime.timedelta(seconds=ArrivalT)
            #--- GroundTime ---
            GroundTime = np.random.beta(2,5)/0.2*BetaMean# np.random.uniform(*TimeRange)
            if GroundTime<TimeRange[0]: GroundTime = TimeRange[0]
            if MinTimeStep is not None: GroundTime = getRoundedThreshold(GroundTime,MinTimeStep)
            #--- Departure ---
            Departure = Arrival + datetime.timedelta(seconds=GroundTime)
            #.. Check Airport Closed / Night Stay ..
            if NightStayMandatory:
                Arrival_s       = (Arrival  -RefDate).total_seconds()
                Departure_s     = (Departure-RefDate).total_seconds()
                Departure_s_rel = Departure_s%(24*3600)
                OverShoot       = Departure_s-Close_s
                # print Departure_s/3600.,Close_s/3600.,OverShoot/3600.
                if OverShoot>0 or Arrival.day!=Departure.day:# or Departure_s_rel<Opened_s:
                    d0 = getRoundedThreshold(Arrival_s,24*3600,SetInt=True)/(24*3600)
                    Departure_s = (d0+1) + (24*3600) + Open_s + OverShoot
                    # print "Departure_s_new",Departure_s/3600.
                    Departure = RefDate+datetime.timedelta(seconds=Departure_s)

            #..............................
            # Add Days
            #..............................
            
            Day = np.random.randint(self.MaxNrDays)
            Arrival  +=datetime.timedelta(days=Day)
            Departure+=datetime.timedelta(days=Day)

            #----------------------------------------
            # Internal Properties
            #----------------------------------------
            
            #--- Nr Passengers ---
            NrPassenger_per = 1-np.random.beta(1,3)#(2,5)
            if NrPassenger_per<0.15:
                NrPassenger_per = 0.15
            NrPassengers = int(AircraftType.MaxNrPassengers*NrPassenger_per)
            # print "NrPassengers:",NrPassengers

            #--- Domestic ---
            if np.random.uniform()<=0.75: Domestic = True
            else:                         Domestic = False 

            #--- Fueling ---
            if np.random.uniform()<=0.75: NeedsFueling = True
            else:                         NeedsFueling = False 

            #--- Preferences ---
            BayPreference  = None
            GatePreference = None
            if np.random.uniform()<=0.125: 

                BayFeasible = False
                NrTries = 0
                while not BayFeasible:
                    NrTries+=1
                    if NrTries>self.Airport.NrBays*2:
                        print "ERROR: Unable to select feasible Prefered Bay for AircraftType: "+AircraftType.Type
                        break

                    BayPreference = np.random.choice([g for g in self.Airport.Gates if not g.Virtual]).Name
                    if BayPreference in self.Airport.Gates_dict.keys():
                        GatePreference = BayPreference

                    #--- Check Feasibility ---
                    bay = self.Airport.Bays_dict[BayPreference]
                    # print bay
                    if AircraftType.Type in bay.CompatibleAircraftTypes:
                        BayFeasible = True
                    else:
                        # print "Infeasible Type-Bay Combination found!"
                        BayPreference  = None
                        GatePreference = None

                if not BayFeasible:
                    continue

            #----------------------------------------
            # Create Aircraft
            #----------------------------------------
            
            a = AircraftType(ID=ID, Arrival=Arrival,Departure=Departure,NrPassengers=NrPassengers)
            a.Airline        = airline
            a.Domestic       = Domestic
            a.NeedsFueling   = NeedsFueling
            a.BayPreference  = BayPreference
            a.GatePreference = GatePreference

            self.Schedule.append(a)

        if Sort:
            self.Schedule.sort(key=lambda x: x.Arrival)#, reverse=True)

    #================================================================================
    # Export
    #================================================================================

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Print
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def PrintSchedule(self):

        for aircraft in self.Schedule:
            aircraft.PrintInfo()

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Excel
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def ExportScheduleToExcel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        #----------------------------------------
        # Write Header
        #----------------------------------------
        
        ws.cell(row=1, column=1 ).value = "ID"
        ws.cell(row=1, column=2 ).value = "Type"
        ws.cell(row=1, column=3 ).value = "Airline"
        ws.cell(row=1, column=4 ).value = "Arrival [datetime]"
        ws.cell(row=1, column=5 ).value = "Departure [datetime]"
        ws.cell(row=1, column=6 ).value = "GroundTime [h]"
        ws.cell(row=1, column=7 ).value = "NrPassengers [#]"
        ws.cell(row=1, column=8 ).value = "Domestic [bool]"
        ws.cell(row=1, column=9 ).value = "NeedsFueling [bool]"
        ws.cell(row=1, column=10).value = "BayPreference"
        ws.cell(row=1, column=11).value = "GatePreference"
        ws.cell(row=1, column=12).value = "NightStay [bool]"

        #----------------------------------------
        # Write Data
        #----------------------------------------

        for i in range(self.MaxNrAircraft):
            aircraft = self.Schedule[i]
            ws.cell(row=i+2, column=1 ).value = aircraft.ID
            ws.cell(row=i+2, column=2 ).value = aircraft.Type
            ws.cell(row=i+2, column=3 ).value = aircraft.Airline.Name
            ws.cell(row=i+2, column=4 ).value = aircraft.Arrival#.time()
            ws.cell(row=i+2, column=5 ).value = aircraft.Departure#.time()
            ws.cell(row=i+2, column=6 ).value = round(aircraft.GroundTime/3600.,1)
            ws.cell(row=i+2, column=7 ).value = aircraft.NrPassengers
            ws.cell(row=i+2, column=8 ).value = aircraft.Domestic
            ws.cell(row=i+2, column=9 ).value = aircraft.NeedsFueling
            ws.cell(row=i+2, column=10).value = aircraft.BayPreference
            ws.cell(row=i+2, column=11).value = aircraft.GatePreference
            ws.cell(row=i+2, column=12).value = aircraft.Arrival.day!=aircraft.Departure.day

        title = self.FormatTitle("Schedule.xlsx")
        wb.save(os.path.join(self.ScheduleFolder,title))
        print ">"+" "+"Exported '"+title+"'"

    #================================================================================
    # Visualize
    #================================================================================

    def Visualize(self,Show=False):
    
        self.ShowGanttChart(          Show=False)
        self.ShowAircraftOnGround(    Show=False)
        self.ShowAircraftGroundTime(  Show=False)
        self.ShowAircraftNrPassengers(Show=False)
        self.ShowAircraftTypes(       Show=False)
        self.ShowAirlines(            Show=False)
        

        if Show:
            plt.show()
        plt.close('all')

    def ShowAirportDayLines(self,MaxNrDays=None):
        if MaxNrDays is None: MaxNrDays = self.MaxNrDays
        for i in range(MaxNrDays):
            Open_t  = (self.Airport.T_Open -dt0).total_seconds()/3600. +24*i
            Close_t = (self.Airport.T_Close-dt0).total_seconds()/3600. +24*i
            plt.axvline(x=Open_t ) # Open line
            plt.axvline(x=Close_t) # Close line
            plt.axvline(x=24*(i+1),color='red') # Day line

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # GanttChart
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def ShowGanttChart(self,Show=True):

        fig,ax=plt.subplots(figsize=(16,9),dpi=self.dpi)

        #----------------------------------------
        # Preprocess
        #----------------------------------------
        
        # MinDate = ReferenceDate
        MinDate = GetDate(np.min([a.Arrival for a in self.Schedule]))

        #----------------------------------------
        # Set Data
        #----------------------------------------
        
        labels=[]
        for i in range(len(self.Schedule)):
            aircraft = self.Schedule[i]
            labels.append(aircraft.ID)
            color = aircraft.GetColor(Mode="GroundTime")
            t_a = (aircraft.Arrival  -MinDate).total_seconds()/3600.
            t_d = (aircraft.Departure-MinDate).total_seconds()/3600.
            dt = t_d-t_a
            ax.broken_barh([(t_a,dt)], (i-0.4,0.8), color=color)

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        self.ShowAirportDayLines()

        #..............................
        # Labels
        #..............................
        
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=5) 
        ax.set_xlabel("Time [h]")

        #..............................
        # Layout
        #..............................
        
        plt.xlim(left=0)
        plt.tight_layout()
        title = self.FormatTitle("Schedule")
        plt.savefig(os.path.join(self.ScheduleFolder,title))
        print "> Saved Figure '"+title+"'"
        if Show: plt.show(title)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Aircraft on Ground
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def ShowAircraftOnGround(self,NrElements=int(24*(60/5.)),AddAircraftTypes=True,Show=True):

        fig,ax=plt.subplots(figsize=(16,9),dpi=self.dpi)

        #----------------------------------------
        # Set Data
        #----------------------------------------
        
        x = []
        y = []
        if AddAircraftTypes: y_Type = OrderedDict()

        I = NrElements*2
        for i in range(I):
            t = float(i)/NrElements * (3600*24)

            NrAircraft = 0
            for aircraft in self.Schedule:
                t_a = aircraft.Arrival_t
                t_d = aircraft.Departure_t
                if t_a<=t and t<=t_d:
                    NrAircraft+=1
                    #--- Track Types ---
                    if AddAircraftTypes:
                        if not aircraft.Type in y_Type.keys():
                            # y_Type[aircraft.Type][i] = 0
                            y_Type[aircraft.Type] = np.zeros(I)
                        y_Type[aircraft.Type][i]+=1
                # print t/3600,t_a/3600,t_d/3600
            # print "t: "+str(round(    t/3600.,2))+" h","NrAircraft",NrAircraft
            x.append(t)
            y.append(NrAircraft)

        x = np.array(x)
        y = np.array(y) 
        NrAircraft_Max = np.max(y)

        color = [[float(na)/NrAircraft_Max,1-float(na)/NrAircraft_Max,0] for na in y]

        x /= 3600.

        plt.bar(x,y,width=24./NrElements,color=color)

        #..............................
        # Aircraft Types
        #..............................
        
        if AddAircraftTypes:
            axes2 = plt.twinx()
            NrTypes = len(y_Type.keys())

            #--- Plots ---
            for i,Type in enumerate(y_Type):
                color = [float(i)/(NrTypes-1)]*3
                y_t = y_Type[Type]

                axes2.plot(x,y_t,color=color)

            #--- Annotate ---
            for i,Type in enumerate(y_Type):
                color = [float(i)/(NrTypes-1)]*3
                y_t = y_Type[Type]

                x_m = x[np.argmax(y_t)]
                y_m = np.max(y_t)
                axes2.annotate(Type, xy=(x_m,y_m), xytext=(x_m+1, y_m+1),
                arrowprops=dict(arrowstyle="simple",facecolor='black',color=color),#, shrink=0.05
                color=color,
                size = 8,
                zorder=100-NrTypes+i
                )

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        self.ShowAirportDayLines()
        
        #..............................
        # Label
        #..............................
        
        plt.xlabel("Time [h]")
        plt.ylabel("Aircraft on ground [#]")

        #..............................
        # Layout
        #..............................
        
        plt.xlim(left=0)
        if 1:
            plt.ylim(  0, np.max(y))
            axes2.set_ylim(0, np.max(y))
        plt.tight_layout()
        title = self.FormatTitle("Aircraft on Ground")
        plt.savefig(os.path.join(self.ScheduleFolder,title))
        print "> Saved Figure '"+title+"'"
        if Show: plt.show(title)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Aircraft GroundTime
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def ShowAircraftGroundTime(self,Show=True):

        fig,ax=plt.subplots(figsize=(16,9),dpi=self.dpi)

        #----------------------------------------
        # Set Data
        #----------------------------------------
 
        GroundTimes = []
        for aircraft in self.Schedule:
            GroundTimes.append(aircraft.GroundTime/3600.)

        db = 0.25#0.5
        bins = np.arange(0,np.max(GroundTimes)+db, db) # None   
        plt.hist(GroundTimes,bins=bins)

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        #..............................
        # Label
        #..............................
        
        plt.xlabel("GroundTime [h]")
        plt.ylabel("Nr of Aircraft [#]")
        
        #..............................
        # Layout
        #..............................
        
        plt.tight_layout()
        title = self.FormatTitle("Aircraft GroundTime")
        plt.savefig(os.path.join(self.ScheduleFolder,title))
        print "> Saved Figure '"+title+"'"
        if Show: plt.show(title)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Aircraft GroundTime
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    def ShowAircraftNrPassengers(self,Show=True):

        fig,ax=plt.subplots(figsize=(16,9),dpi=self.dpi)

        #----------------------------------------
        # Set Data
        #----------------------------------------
 
        NrPassengers = []
        for aircraft in self.Schedule:
            NrPassengers.append(aircraft.NrPassengers)

        db = 25#50
        bins = np.arange(0,np.max(NrPassengers)+db, db) # None   
        plt.hist(NrPassengers,bins=bins)

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        #..............................
        # Label
        #..............................
        
        plt.xlabel("NrPassengers [#]")
        plt.ylabel("Nr of Aircraft [#]")
        
        #..............................
        # Layout
        #..............................
        
        plt.tight_layout()
        title = self.FormatTitle("Aircraft NrPassengers")
        plt.savefig(os.path.join(self.ScheduleFolder,title))
        print "> Saved Figure '"+title+"'"
        if Show: plt.show(title)

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Aircraft Types Histogram
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def ShowAircraftTypes(self,Show=True):

        fig,ax=plt.subplots(figsize=(16,9),dpi=self.dpi)

        #----------------------------------------
        # Set Data
        #----------------------------------------
 
        ACTypes = []
        for aircraft in self.Schedule:
            ACTypes.append(aircraft.Type)

        ACT_Counter = Counter(ACTypes)

        NR_ofAC = ACT_Counter.values()
        AC_model= ACT_Counter.keys()
        
        types  = np.arange(len(ACT_Counter))
        ax.bar(types, NR_ofAC, align='center')

        ax.xaxis.set_major_locator(plt.FixedLocator(types))
        ax.xaxis.set_major_formatter(plt.FixedFormatter(AC_model))

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        #..............................
        # Label
        #..............................
        
        plt.xlabel("Aircraft Types")
        plt.ylabel("Nr of Aircraft [#]")
        
        #..............................
        # Layout
        #..............................
        
        plt.tight_layout()
        title = self.FormatTitle("Aircraft Types")
        plt.savefig(os.path.join(self.ScheduleFolder,title))
        print "> Saved Figure '"+title+"'"
        if Show: plt.show(title)


    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Airlines Histogram
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def ShowAirlines(self,Show=True):

        fig,ax=plt.subplots(figsize=(16,9),dpi=120)

        #----------------------------------------
        # Set Data
        #----------------------------------------
 
        AirlinesHist = []
        for aircraft in self.Schedule:
            AirlinesHist.append(aircraft.Airline.Name)

        Airline_Counter = Counter(AirlinesHist)

        NR_ofAL     = Airline_Counter.values()
        Airlinename = Airline_Counter.keys()
        
        ALnames  = np.arange(len(Airline_Counter))
        ax.bar(ALnames, NR_ofAL, align='center')

        ax.xaxis.set_major_locator(plt.FixedLocator(ALnames))
        ax.xaxis.set_major_formatter(plt.FixedFormatter(Airlinename))

        #----------------------------------------
        # Configure Plot
        #----------------------------------------
        
        #..............................
        # Label
        #..............................
        
        plt.xlabel("Airline")
        plt.ylabel("Nr of Aircraft [#]")
        
        #..............................
        # Layout
        #..............................
        
        plt.tight_layout()
        title = self.FormatTitle("Airlines")
        plt.savefig(os.path.join(self.ScheduleFolder,title))
        print "> Saved Figure '"+title+"'"
        if Show: plt.show(title)




    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Aircraft GroundTime
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def FormatTitle(self,title):
        Seed = np.random.seed_backup #np.random.get_state()[1][0]
        return self.Airport.Name+" ("+str(Seed)+")"+" - "+title

#****************************************************************************************************
# Test Code
#****************************************************************************************************

if __name__=="__main__":
     
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Imports
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    from Airport import Airport

    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Run Test
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    
    airport = Airport()
    airport.PrintInfo()

    SC = ScheduleCreator(airport)
